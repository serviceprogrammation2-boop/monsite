import os
import tempfile
from itertools import groupby
from operator import attrgetter

from django.core.paginator import Paginator
from django.http import HttpResponse
from django.shortcuts import render
from django.template.loader import render_to_string
from weasyprint import HTML

from ..models import Ligne

MAP_SORTIE = {
    1: "Dépôt Ben Arous",
    2: "Gare Routière Nord",
    3: "Gare Routière Sud",
    4: "Convention",
}


def ligne_list(request):
    lignes = Ligne.objects.all()

    code = request.GET.get("code") or request.POST.get("code")
    agence = request.GET.get("agence") or request.POST.get("agence")
    actif = request.GET.get("actif") or request.POST.get("actif")
    sortie = request.GET.get("sortie") or request.POST.get("sortie")

    if agence:
        lignes = lignes.filter(agence=agence)
    if actif in ["1", "true", "True"]:
        lignes = lignes.filter(actif=1)
    elif actif in ["0", "false", "False"]:
        lignes = lignes.filter(actif=0)
    if code:
        lignes = lignes.filter(code__icontains=code)
    if sortie:
        lignes = lignes.filter(sortie=sortie)

    lignes = lignes.order_by("sortie", "ord")

    paginator = Paginator(lignes, 15)
    page_obj = paginator.get_page(request.GET.get("page"))

    return render(request, "blog/ligne_list.html", {
        "MAP_SORTIE": MAP_SORTIE,
        "sortie": sortie or "",
        "code": code or "",
        "agence": agence or "",
        "actif": actif or "",
        "actif_true": actif in ["1", "true", "True"],
        "actif_false": actif in ["0", "false", "False"],
        "page_obj": page_obj,
        "lignes": page_obj.object_list,
    })


def ligne_pdf(request):
    lignes = Ligne.objects.all()

    code = request.GET.get("code")
    agence = request.GET.get("agence")
    actif = request.GET.get("actif")

    if agence:
        lignes = lignes.filter(agence=agence)
    if actif in ["1", "true", "True"]:
        lignes = lignes.filter(actif=1)
    elif actif in ["0", "false", "False"]:
        lignes = lignes.filter(actif=0)
    if code:
        lignes = lignes.filter(code__icontains=code)

    lignes = lignes.order_by("sortie", "ord")

    groupes = {}
    for sortie, items in groupby(lignes, key=attrgetter("sortie")):
        groupes[sortie] = list(items)

    html_string = render_to_string("blog/ligne_pdf.html", {
        "groupes": groupes,
        "map_sortie": MAP_SORTIE,
    })

    response = HttpResponse(content_type="application/pdf")
    response['Content-Disposition'] = 'inline; filename="lignes.pdf"'

    tmp_path = os.path.join(tempfile.gettempdir(), "lignes_export.pdf")
    HTML(string=html_string).write_pdf(target=tmp_path)

    with open(tmp_path, 'rb') as f:
        response.write(f.read())
    os.remove(tmp_path)
    return response
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from django.http import HttpResponse

from itertools import groupby
from operator import attrgetter


def ligne_excel(request):
    lignes = Ligne.objects.all()

    code = request.GET.get("code")
    agence = request.GET.get("agence")
    actif = request.GET.get("actif")

    if agence:
        lignes = lignes.filter(agence=agence)
    if actif in ["1", "true", "True"]:
        lignes = lignes.filter(actif=1)
    elif actif in ["0", "false", "False"]:
        lignes = lignes.filter(actif=0)
    if code:
        lignes = lignes.filter(code__icontains=code)

    lignes = lignes.order_by("sortie", "ord")

    # Créer le workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Lignes"

    # Styles
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="2E75B6", end_color="2E75B6", fill_type="solid")
    group_font = Font(bold=True, color="FFFFFF", size=10)
    group_fill = PatternFill(start_color="5B9BD5", end_color="5B9BD5", fill_type="solid")
    center = Alignment(horizontal="center", vertical="center")
    thin = Side(style="thin", color="CCCCCC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # En-têtes des colonnes
    headers = ["Code", "Agence", "Sortie", "Ordre", "Actif"]
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center
        cell.border = border

    ws.row_dimensions[1].height = 25

    # Largeur des colonnes
    ws.column_dimensions["A"].width = 20
    ws.column_dimensions["B"].width = 20
    ws.column_dimensions["C"].width = 25
    ws.column_dimensions["D"].width = 10
    ws.column_dimensions["E"].width = 10

    row = 2
    for sortie, items in groupby(lignes, key=attrgetter("sortie")):
        # Ligne de groupe
        label = MAP_SORTIE.get(sortie, f"Sortie {sortie}")
        merge_end = chr(ord("A") + len(headers) - 1)
        ws.merge_cells(f"A{row}:{merge_end}{row}")
        cell = ws.cell(row=row, column=1, value=label)
        cell.font = group_font
        cell.fill = group_fill
        cell.alignment = center
        cell.border = border
        ws.row_dimensions[row].height = 20
        row += 1

        # Lignes de données
        for ligne in items:
            data = [
                ligne.code,
                ligne.agence,
                MAP_SORTIE.get(ligne.sortie, ligne.sortie),
                ligne.ord,
                "Oui" if ligne.actif else "Non",
            ]
            for col, value in enumerate(data, start=1):
                cell = ws.cell(row=row, column=col, value=value)
                cell.alignment = center
                cell.border = border
                # Zebra striping
                if row % 2 == 0:
                    cell.fill = PatternFill(start_color="EBF3FB", end_color="EBF3FB", fill_type="solid")
            row += 1

    # Retourner le fichier
    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = 'attachment; filename="lignes.xlsx"'
    wb.save(response)
    return response